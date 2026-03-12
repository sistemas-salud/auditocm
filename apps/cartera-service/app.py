#Prueba 1
# from flask import Flask

# app = Flask(__name__)

# @app.route("/")
# def home():
#     return "Modulo Cartera funcionando 🚀"

# if __name__ == "__main__":
#     app.run(host="127.0.0.1", port=5050, debug=True)

#Prueba 2
# from flask import Flask, jsonify
# from db import get_connection

# app = Flask(__name__)

# @app.route("/")
# def home():
#     return "Modulo Cartera funcionando 🚀"

# @app.route("/test-db")
# def test_db():
#     try:
#         conn = get_connection()
#         cursor = conn.cursor()
#         cursor.execute("SELECT NOW();")
#         result = cursor.fetchone()
#         cursor.close()
#         conn.close()

#         return jsonify({
#             "status": "success",
#             "server_time": result[0]
#         })
#     except Exception as e:
#         return jsonify({
#             "status": "error",
#             "message": str(e)
#         })

# if __name__ == "__main__":
#     app.run(host="127.0.0.1", port=5050, debug=True)

#Prueba 3
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from flask import Flask, request, jsonify, render_template
from flask import render_template
from db import get_connection
from flask import Blueprint



app = Flask(__name__)

# ==========================================================
# 🔹 INDEX
# ==========================================================

@app.route("/")
def index():
    return render_template("index.html")


# ==========================================================
# 🔹 VISTA PRINCIPAL CARTERA
# ==========================================================

@app.route("/cartera")
def cartera():
    # Obtener filtros existentes
    nit = request.args.get("nit", "").strip()
    acta = request.args.get("acta", "").strip()
    year = request.args.get("year", "").strip()
    
    # 🔥 NUEVOS FILTROS
    no_factura = request.args.get("no_factura", "").strip()
    fecha_factura = request.args.get("fecha_factura", "").strip()
    fecha_prestacion = request.args.get("fecha_prestacion", "").strip()
    fecha_radicacion = request.args.get("fecha_radicacion", "").strip()
    fecha_pago = request.args.get("fecha_pago", "").strip()
    observacion = request.args.get("observacion", "").strip()
    
    print("="*50)
    print("FILTROS RECIBIDOS:")
    print(f"NIT: '{nit}'")
    print(f"Acta: '{acta}'")
    print(f"Año: '{year}'")
    print(f"No Factura: '{no_factura}'")
    print(f"Fecha Factura: '{fecha_factura}'")
    print(f"Fecha Prestación: '{fecha_prestacion}'")
    print(f"Fecha Radicación: '{fecha_radicacion}'")
    print(f"Fecha Pago: '{fecha_pago}'")
    print(f"Observación: '{observacion}'")
    print("="*50)
    
    data = []
    totales = {
        "VALOR_FACTURA": 0,
        "VALOR_GLOSADO": 0,
        "VALOR_RECONOCIDO": 0,
        "VALOR_PAGADO": 0,
        "VALOR_SALDO": 0,
        "VALOR_CONTINGENTE": 0,
        "CUENTA_NO_FACTURAS": 0
    }
    
    # Verificar si hay algún filtro activo
    filtros_activos = [nit, acta, year, no_factura, fecha_factura, 
                       fecha_prestacion, fecha_radicacion, fecha_pago, observacion]
    
    if any(filtros_activos):
        try:
            conn = get_connection()
            cursor = conn.cursor(dictionary=True)
            
            # Consulta base
            query = """
                SELECT 
                    'MIG' AS base_datos,
                    fm.NIT AS nit,
                    fm.IPS AS ips,
                    fm.ACTA AS acta,
                    YEAR(fm.FECHA_FACTURA) AS año_factura,
                    fm.NO_FACTURA AS numero_factura,
                    fm.FECHA_FACTURA,
                    fm.FECHA_PRESTACION,
                    fm.FECHA_RADICACION,
                    fm.FECHA_PAGO,
                    fm.OBSERVACION_PAGO,
                    fm.VALOR_FACTURA,
                    fm.VALOR_GLOSADO,
                    fm.VALOR_RECONOCIDO,
                    fm.VALOR_PAGADO,
                    fm.VALOR_SALDO,
                    fm.PASIVO_CONTINGENTE AS VALOR_CONTINGENTE
                FROM facturas_mig fm
                WHERE 1=1
            """
            
            params = []
            
            # Filtros existentes
            if nit:
                query += " AND fm.NIT = %s"
                params.append(nit)
            
            if acta:
                query += " AND fm.ACTA = %s"
                params.append(acta)
            
            if year:
                try:
                    year_int = int(year)
                    query += " AND YEAR(fm.FECHA_FACTURA) = %s"
                    params.append(year_int)
                except ValueError:
                    pass
            
            # 🔥 NUEVOS FILTROS
            if no_factura:
                query += " AND fm.NO_FACTURA LIKE %s"
                params.append(f"%{no_factura}%")
            
            if fecha_factura:
                query += " AND DATE(fm.FECHA_FACTURA) = %s"
                params.append(fecha_factura)
            
            if fecha_prestacion:
                query += " AND DATE(fm.FECHA_PRESTACION) = %s"
                params.append(fecha_prestacion)
            
            if fecha_radicacion:
                query += " AND DATE(fm.FECHA_RADICACION) = %s"
                params.append(fecha_radicacion)
            
            if fecha_pago:
                query += " AND DATE(fm.FECHA_PAGO) = %s"
                params.append(fecha_pago)
            
            if observacion:
                query += " AND fm.OBSERVACION_PAGO LIKE %s"
                params.append(f"%{observacion}%")
            
            # Ordenar resultados
            query += " ORDER BY fm.FECHA_FACTURA DESC, fm.NO_FACTURA"
            
            print(f"Query: {query}")
            print(f"Parámetros: {params}")
            
            cursor.execute(query, tuple(params))
            data = cursor.fetchall()
            
            print(f"Registros encontrados: {len(data)}")
            
            # Calcular totales
            for row in data:
                totales["VALOR_FACTURA"] += float(row["VALOR_FACTURA"] or 0)
                totales["VALOR_GLOSADO"] += float(row["VALOR_GLOSADO"] or 0)
                totales["VALOR_RECONOCIDO"] += float(row["VALOR_RECONOCIDO"] or 0)
                totales["VALOR_PAGADO"] += float(row["VALOR_PAGADO"] or 0)
                totales["VALOR_SALDO"] += float(row["VALOR_SALDO"] or 0)
                totales["VALOR_CONTINGENTE"] += float(row["VALOR_CONTINGENTE"] or 0)
                totales["CUENTA_NO_FACTURAS"] += 1
            
            cursor.close()
            conn.close()
            
        except Exception as e:
            print(f"Error en consulta: {str(e)}")
            import traceback
            traceback.print_exc()
            return f"Error: {str(e)}", 500
    
    return render_template(
        "cartera.html",
        cartera=data,
        totales=totales,
        filtros={
            "nit": nit,
            "acta": acta,
            "year": year,
            "no_factura": no_factura,
            "fecha_factura": fecha_factura,
            "fecha_prestacion": fecha_prestacion,
            "fecha_radicacion": fecha_radicacion,
            "fecha_pago": fecha_pago,
            "observacion": observacion
        }
    )



@app.template_filter('formato_numero')
def formato_numero(value):
    """Convierte a float y formatea número"""
    try:
        return "{:,.0f}".format(float(value or 0))
    except (ValueError, TypeError):
        return "0"
# ==========================================================
# 🔹 MODULOS DE CARGUE
# ==========================================================

@app.route("/upload/migrantes")
def upload_migrantes():
    return render_template("upload_migrantes.html")

@app.route("/upload/urgencias")
def upload_urgencias():
    return render_template("upload_urgencias.html")

@app.route("/upload/nopos")
def upload_nopos():
    return render_template("upload_nopos.html")


@app.route("/cartera/acta")
def cartera_por_acta():
    try:
        nit = request.args.get("nit")
        year = request.args.get("year")
        acta = request.args.get("acta")
        no_factura = request.args.get("no_factura")

        if not nit and not acta and not no_factura:
            return jsonify({
                "status": "error",
                "message": "Debe ingresar al menos un filtro"
            }), 400

        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        params = []

        # 🔥 Determinar campo de agrupación
        if no_factura:
            campo_group = "fm.no_factura"
            campo_select = "fm.no_factura AS acta"
        else:
            campo_group = "fm.acta"
            campo_select = "fm.acta"

        query = f"""
            SELECT 
                {campo_select},
                SUM(fm.VALOR_FACTURA) AS SUMA_VALOR_FACTURA,
                SUM(fm.VALOR_GLOSADO) AS SUMA_DE_VALOR_GLOSADO,
                SUM(fm.VALOR_RECONOCIDO) AS SUMA_DE_VALOR_RECONOCIDO,
                SUM(fm.VALOR_PAGADO) AS SUMA_DE_VALOR_PAGADO,
                SUM(fm.VALOR_SALDO) AS SUMA_DE_VALOR_SALDO,
                SUM(fm.PASIVO_CONTINGENTE) AS SUMA_DE_PASIVO_CONTINGENTE,
                COUNT(fm.no_factura) AS CUENTA_No_FACTURA
            FROM facturas_mig fm
            WHERE 1=1
        """

        if nit:
            query += " AND fm.nit = %s"
            params.append(nit)

        if acta:
            query += " AND fm.acta = %s"
            params.append(acta)

        if no_factura:
            query += " AND fm.no_factura = %s"
            params.append(no_factura)

        if year and year != "*":
            query += " AND YEAR(fm.fecha_factura) = %s"
            params.append(year)

        query += f" GROUP BY {campo_group}"

        cursor.execute(query, tuple(params))
        result = cursor.fetchall()

        cursor.close()
        conn.close()

        return jsonify({
            "status": "success",
            "data": result
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

# ==========================================================
# 🔹 RESUMEN GENERAL (SIN FILTROS)
# ==========================================================

@app.route("/cartera/resumen")
def cartera_resumen():
    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        query = """
            SELECT 
                estado,
                COALESCE(SUM(valor),0) AS total_valor,
                COALESCE(SUM(cantidad),0) AS total_cantidad
            FROM (
                SELECT 
                    'EN PROCESO DE CONCILIACION' AS estado,
                    SUM(CASE WHEN acta = 'EN PROCESO DE CONCILIACION' THEN VALOR_FACTURA ELSE 0 END) AS valor,
                    COUNT(CASE WHEN acta = 'EN PROCESO DE CONCILIACION' THEN 1 END) AS cantidad
                FROM facturas_mig

                UNION ALL

                SELECT 
                    'EN PROCESO DE AUDITORIA',
                    SUM(CASE WHEN acta = 'EN PROCESO DE AUDITORIA' THEN VALOR_FACTURA ELSE 0 END),
                    COUNT(CASE WHEN acta = 'EN PROCESO DE AUDITORIA' THEN 1 END)
                FROM facturas_mig

                UNION ALL

                SELECT 
                    'SIN ASIGNAR',
                    SUM(CASE WHEN acta = 'SIN ASIGNAR' THEN VALOR_FACTURA ELSE 0 END),
                    COUNT(CASE WHEN acta = 'SIN ASIGNAR' THEN 1 END)
                FROM facturas_mig
            ) t
            GROUP BY estado
        """

        cursor.execute(query)
        result = cursor.fetchall()

        cursor.close()
        conn.close()

        return jsonify({
            "status": "success",
            "data": result
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500


# ==========================================================
# 🔹 AUTOCOMPLETE NIT (CON NOMBRE IPS)
# ==========================================================

@app.route("/autocomplete_nit")
def autocomplete_nit():
    term = request.args.get("term", "")
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    
    if term:
        # Si hay término de búsqueda, filtrar
        cursor.execute("""
            SELECT DISTINCT 
                fm.NIT AS nit,
                fm.IPS AS nombre
            FROM facturas_mig fm
            WHERE fm.NIT LIKE %s OR fm.IPS LIKE %s
            LIMIT 10
        """, (f"{term}%", f"%{term}%"))
    else:
        # Si no hay término, mostrar algunos ejemplos
        cursor.execute("""
            SELECT DISTINCT 
                fm.NIT AS nit,
                fm.IPS AS nombre
            FROM facturas_mig fm
            WHERE fm.NIT IS NOT NULL AND fm.NIT != ''
            LIMIT 10
        """)
    
    resultados = cursor.fetchall()
    cursor.close()
    conn.close()
    
    return jsonify(resultados)


# ==========================================================
# 🔹 AUTOCOMPLETE ACTA
# ==========================================================

@app.route("/autocomplete_acta")
def autocomplete_acta():
    term = request.args.get("term", "")
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    
    if term:
        cursor.execute("""
            SELECT DISTINCT fm.acta AS acta,
                   YEAR(fm.fecha_factura) as año
            FROM facturas_mig fm
            WHERE fm.acta LIKE %s
            LIMIT 10
        """, (f"{term}%",))
    else:
        cursor.execute("""
            SELECT DISTINCT fm.acta AS acta,
                   YEAR(fm.fecha_factura) as año
            FROM facturas_mig fm
            WHERE fm.acta IS NOT NULL AND fm.acta != ''
            LIMIT 10
        """)
    
    resultados = cursor.fetchall()
    cursor.close()
    conn.close()
    
    return jsonify(resultados)

# ==========================================================
# 🔹 AUTOCOMPLETE NO_FACTURA
# ==========================================================

@app.route("/autocomplete_factura")
def autocomplete_factura():
    term = request.args.get("term", "")
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    if term:
        cursor.execute("""
            SELECT DISTINCT fm.no_factura AS No_factura,
                   YEAR(fm.fecha_factura) as año
            FROM facturas_mig fm
            WHERE fm.no_factura LIKE %s
            LIMIT 10
        """, (f"{term}%",))
    else:
        cursor.execute("""
            SELECT DISTINCT fm.no_factura AS NO_factura,
                   YEAR(fm.fecha_factura) as año
            FROM facturas_mig fm
            WHERE fm.no_factura IS NOT NULL AND fm.no_factura != ''
            LIMIT 10
        """)

    resultados = cursor.fetchall()
    cursor.close()
    conn.close()

    return jsonify(resultados)


# ==========================================================
# 🔹 AUTOCOMPLETE OBSERVACION DE PAGO
# ==========================================================

@app.route("/autocomplete_observacion")
def autocomplete_observacion():
    term = request.args.get("term", "")
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("""
        SELECT DISTINCT fm.OBSERVACION_PAGO AS observacion
        FROM facturas_mig fm
        WHERE fm.OBSERVACION_PAGO IS NOT NULL 
          AND fm.OBSERVACION_PAGO != ''
          AND fm.OBSERVACION_PAGO LIKE %s
        LIMIT 10
    """, (f"%{term}%",))
    
    resultados = cursor.fetchall()
    cursor.close()
    conn.close()
    
    return jsonify(resultados)


# ==========================================================
# 🔹 DETALLE DE FACTURAS POR ACTA
# ==========================================================

@app.route("/cartera/detalle")
def cartera_detalle():
    try:
        acta = request.args.get("acta")
        nit = request.args.get("nit")
        no_factura = request.args.get("no_factura")
        
        # Debug
        print(f"Detalle solicitado - acta: {acta}, nit: {nit}, factura: {no_factura}")
        
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)
        
        query = """
            SELECT 
                fm.no_factura,
                fm.acta,
                fm.nit,
                fm.ips,
                fm.VALOR_FACTURA,
                fm.VALOR_GLOSADO,
                fm.VALOR_RECONOCIDO,
                fm.VALOR_PAGADO,
                fm.VALOR_SALDO
            FROM facturas_mig fm
            WHERE 1=1
        """
        
        params = []
        
        if no_factura:
            query += " AND fm.no_factura = %s"
            params.append(no_factura)
            print(f"Filtrando por factura: {no_factura}")
        elif acta:
            query += " AND fm.acta = %s"
            params.append(acta)
            print(f"Filtrando por acta: {acta}")
        elif nit:
            query += " AND fm.nit = %s"
            params.append(nit)
            print(f"Filtrando por nit: {nit}")
        else:
            return jsonify({
                "status": "error",
                "message": "Debe proporcionar al menos un filtro (acta, nit o no_factura)"
            }), 400
        
        query += " ORDER BY fm.no_factura"
        
        print(f"Query detalle: {query}")
        print(f"Parámetros detalle: {params}")
        
        cursor.execute(query, tuple(params))
        result = cursor.fetchall()
        
        print(f"Registros encontrados en detalle: {len(result)}")
        
        cursor.close()
        conn.close()
        
        return jsonify({
            "status": "success",
            "data": result
        })
        
    except Exception as e:
        print(f"Error en detalle: {str(e)}")
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500



# ==========================================================
# 🔹 DESCARGAR EN EXCEL
# ==========================================================

from flask import send_file
import pandas as pd
import io
from datetime import datetime
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter

@app.route("/cartera/detalle_excel")
def cartera_detalle_excel():
    try:
        acta = request.args.get("acta")
        no_factura = request.args.get("no_factura")

        conn = get_connection()

        query = """
            SELECT 
                fm.no_factura AS 'No Factura',
                fm.acta AS 'Acta',
                fm.nit AS 'NIT',
                fm.ips AS 'IPS',
                fm.VALOR_FACTURA AS 'Valor Factura',
                fm.VALOR_GLOSADO AS 'Valor Glosado',
                fm.VALOR_RECONOCIDO AS 'Valor Reconocido',
                fm.VALOR_PAGADO AS 'Valor Pagado',
                fm.VALOR_SALDO AS 'Valor Saldo'
            FROM facturas_mig fm
            WHERE 1=1
        """

        params = []

        if no_factura:
            query += " AND fm.no_factura = %s"
            params.append(no_factura)

        elif acta:
            query += " AND fm.acta = %s"
            params.append(acta)

        df = pd.read_sql(query, conn, params=params)
        conn.close()

        # 🔥 Agregar fila de totales
        totales = df.select_dtypes(include='number').sum()
        fila_total = {col: '' for col in df.columns}
        fila_total['No Factura'] = 'TOTAL'
        for col in totales.index:
            fila_total[col] = totales[col]

        df = pd.concat([df, pd.DataFrame([fila_total])], ignore_index=True)

        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Detalle')

            workbook = writer.book
            sheet = writer.sheets['Detalle']

            # 🔥 Formato encabezado
            for cell in sheet[1]:
                cell.font = Font(bold=True)

            # 🔥 Formato moneda
            columnas_monetarias = [
                'Valor Factura',
                'Valor Glosado',
                'Valor Reconocido',
                'Valor Pagado',
                'Valor Saldo'
            ]

            for col_name in columnas_monetarias:
                col_index = df.columns.get_loc(col_name) + 1
                col_letter = get_column_letter(col_index)

                for row in range(2, sheet.max_row + 1):
                    sheet[f"{col_letter}{row}"].number_format = '"$"#,##0.00'

            # 🔥 Negrita en fila TOTAL
            ultima_fila = sheet.max_row
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=ultima_fila, column=col).font = Font(bold=True)

            # 🔥 Ajustar ancho columnas
            for col in sheet.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                sheet.column_dimensions[col_letter].width = max_length + 2

            # 🔥 Agregar fecha de generación
            sheet["A1"].alignment = Alignment(horizontal="left")

            sheet.insert_rows(1)
            sheet["A1"] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            sheet["A1"].font = Font(italic=True)

        output.seek(0)

        # 🔥 Nombre dinámico
        nombre_archivo = "detalle_facturas.xlsx"

        if acta:
            nombre_archivo = f"Acta_{acta}.xlsx"

        if no_factura:
            nombre_archivo = f"Factura_{no_factura}.xlsx"

        return send_file(
            output,
            download_name=nombre_archivo,
            as_attachment=True
        )

    except Exception as e:
        return str(e), 500

# ==========================================================
# 🔹 DEBUG PARA VER LOS VALORES DISPONIBLES 
# ==========================================================

@app.route("/debug/valores")
def debug_valores():
    """Endpoint temporal para ver qué valores existen en la BD"""
    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)
        
        resultados = {}
        
        # Ver los primeros 10 NITs diferentes
        cursor.execute("""
            SELECT DISTINCT nit, ips 
            FROM facturas_mig 
            WHERE nit IS NOT NULL AND nit != ''
            LIMIT 10
        """)
        resultados['nits'] = cursor.fetchall()
        
        # Ver los primeros 10 Actas diferentes
        cursor.execute("""
            SELECT DISTINCT acta 
            FROM facturas_mig 
            WHERE acta IS NOT NULL AND acta != ''
            LIMIT 10
        """)
        resultados['actas'] = cursor.fetchall()
        
        # Ver los años disponibles
        cursor.execute("""
            SELECT DISTINCT YEAR(fecha_factura) as año
            FROM facturas_mig 
            WHERE fecha_factura IS NOT NULL
            ORDER BY año
        """)
        resultados['años'] = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return jsonify({
            "total_registros": 210264,
            "muestras": resultados
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ==========================================================
# 🔹 RUN LOCAL
# ==========================================================

if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5050, debug=True)

